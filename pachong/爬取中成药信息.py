import openpyxl
from bs4 import BeautifulSoup
import requests


def extract_info(text):
    sections = text.split('【')

    # 注释掉原来的变量
    # function = ''
    # dosage = ''
    # storage = ''
    taste = ''
    channel = ''

    for section in sections:
        if section.startswith("性味与归经】"):
            content = section.replace("性味与归经】", "").strip()
            parts = content.split('。', 1)  # 使用第一个句号分割
            if len(parts) == 2:
                taste, channel = parts
            else:
                taste = parts[0]  # 如果没有句号，则整个内容都视为性味

    return taste, channel


def get_drug_info(drug_name):
    if not drug_name:
        return None

    base_url = "https://db.ouryao.com/yd2020/index.php?k="
    search_url = base_url + drug_name
    print(f"Drug name: {drug_name}")
    print(f"Search URL: {search_url}")
    response = requests.get(search_url)

    soup = BeautifulSoup(response.content, 'html.parser')
    drug_link = soup.find('a', string=drug_name)
    if not drug_link:
        return None

    detail_url = "https://db.ouryao.com/yd2020/" + drug_link['href']
    detail_response = requests.get(detail_url)
    detail_soup = BeautifulSoup(detail_response.content, 'html.parser')
    info = detail_soup.find('pre', {'id': 'content_text'})

    if not info:
        return None

    return extract_info(info.text)


def main():
    workbook = openpyxl.load_workbook('1.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        for cell in row:
            drug_name = cell.value
            info = get_drug_info(drug_name)
            if info:
                taste, channel = info
                cell.offset(column=4).value = taste
                cell.offset(column=5).value = channel

    workbook.save('1.xlsx')


if __name__ == "__main__":
    main()
